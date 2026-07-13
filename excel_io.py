"""Excel import and export for the Family Investment Tracker."""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from typing import List, Dict, Any, Optional

import models
import metrics as m

# ── Column name mapping for import ───────────────────────────────────────────

_COL_MAP: Dict[str, List[str]] = {
    'company':              ['company', 'company name', 'name'],
    'round':                ['round', 'round name', 'round type', 'series', 'funding round'],
    'date':                 ['date', 'investment date', 'round date', 'close date'],
    'amount_invested':      ['amount invested', 'amount', 'invested', 'our investment',
                             'investment amount', 'investment'],
    'pre_money_valuation':  ['pre-money', 'pre money', 'pre_money', 'pre-money valuation',
                             'pre money valuation', 'premoney'],
    'post_money_valuation': ['post-money', 'post money', 'post_money', 'post-money valuation',
                             'post money valuation', 'postmoney'],
    'shares':               ['shares', 'shares received', 'number of shares', 'our shares'],
    'price_per_share':      ['price per share', 'share price', 'pps', 'price/share'],
    'ownership_pct':        ['ownership', 'ownership %', 'ownership%', 'our ownership',
                             '% ownership', 'stake %', 'stake'],
    'current_valuation':    ['current valuation', 'valuation', 'latest valuation',
                             'current company valuation'],
}


def _map_header(header: str):
    h = str(header).strip().lower()
    for field, variants in _COL_MAP.items():
        if h in variants:
            return field
    return None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return None


def _to_date(v) -> str:
    if v is None:
        return ''
    if isinstance(v, (date, datetime)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s[:10]


# ── Import ────────────────────────────────────────────────────────────────────

def parse_excel(path: str) -> List[Dict[str, Any]]:
    """Read an Excel file and return a list of normalised row dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    # Map header → field name
    header_row = rows[0]
    col_idx: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h:
            field = _map_header(str(h))
            if field and field not in col_idx:
                col_idx[field] = i

    if 'company' not in col_idx:
        raise ValueError(
            "Could not find a 'Company' column in the spreadsheet.\n"
            "Please make sure the first row contains column headers."
        )

    result = []
    for row in rows[1:]:
        if not any(cell is not None for cell in row):
            continue

        def get(field):
            i = col_idx.get(field)
            return row[i] if (i is not None and i < len(row)) else None

        company = get('company')
        if not company:
            continue

        result.append({
            'company':              str(company).strip(),
            'round':                str(get('round') or '').strip(),
            'date':                 _to_date(get('date')),
            'amount_invested':      _to_float(get('amount_invested')),
            'pre_money_valuation':  _to_float(get('pre_money_valuation')),
            'post_money_valuation': _to_float(get('post_money_valuation')),
            'shares':               _to_float(get('shares')),
            'price_per_share':      _to_float(get('price_per_share')),
            'ownership_pct':        _to_float(get('ownership_pct')),
            'current_valuation':    _to_float(get('current_valuation')),
        })

    return result


def import_rows(rows: List[Dict[str, Any]], on_conflict: str = 'update') -> None:
    """Write parsed rows to the database."""
    for row in rows:
        # Find or create company
        all_companies = models.get_all_companies()
        company = next(
            (c for c in all_companies if c['name'].lower() == row['company'].lower()),
            None
        )
        if company:
            cid = company['id']
            if row.get('current_valuation'):
                models.update_company(cid, current_valuation=row['current_valuation'])
        else:
            cid = models.add_company(
                name=row['company'],
                current_valuation=row.get('current_valuation'),
            )

        # Skip if no round data at all
        if not row.get('round') and not row.get('amount_invested'):
            continue

        # Find or create round
        existing_rounds = models.get_rounds(cid)
        match = next(
            (r for r in existing_rounds
             if r['round_name'].lower() == (row['round'] or '').lower()
             and r['date'] == row['date']),
            None
        )

        if match:
            if on_conflict == 'update':
                models.update_round(match['id'],
                    round_name=row['round'] or match['round_name'],
                    date=row['date'] or match['date'],
                    amount_invested=row['amount_invested'] or match['amount_invested'],
                    pre_money_valuation=row['pre_money_valuation'],
                    post_money_valuation=row['post_money_valuation'],
                    shares_received=row['shares'],
                    price_per_share=row['price_per_share'],
                    ownership_pct=row['ownership_pct'],
                )
            # else: skip
        else:
            models.add_round(
                company_id=cid,
                round_name=row['round'] or 'Round',
                date=row['date'],
                amount_invested=row['amount_invested'],
                pre_money_valuation=row['pre_money_valuation'],
                post_money_valuation=row['post_money_valuation'],
                shares_received=row['shares'],
                price_per_share=row['price_per_share'],
                ownership_pct=row['ownership_pct'],
            )


# ── Export ────────────────────────────────────────────────────────────────────

_HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
_HDR_FILL  = PatternFill(fill_type='solid', fgColor='1E293B')
_GREEN_FILL = PatternFill(fill_type='solid', fgColor='DCFCE7')
_RED_FILL   = PatternFill(fill_type='solid', fgColor='FEE2E2')
_CENTER     = Alignment(horizontal='center')
_CURRENCY   = '#,##0.00'
_PCT        = '0.00"%"'


def export_portfolio(path: str) -> None:
    sym       = models.get_setting('currency', '$')
    companies = models.get_all_companies()
    wb        = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Portfolio Summary"

    headers = ["Company", "Sector", "Country", "Round", "Date",
               "Amount Invested", "Pre-Money Val.", "Post-Money Val.",
               "Shares", "Price/Share", "Ownership %",
               "Current Co. Valuation", "Our Stake Value",
               "Gain / Loss", "ROI %", "MOIC", "Status"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER

    ws.freeze_panes = 'A2'
    row_num = 2

    for c in companies:
        rounds = models.get_rounds(c['id'])
        met    = m.company_metrics(rounds, c.get('current_valuation'))

        for r in rounds:
            invested     = r.get('amount_invested') or 0
            total_inv    = met['total_invested'] or 1
            proportion   = invested / total_inv if total_inv else 0
            stake_value  = (met['current_value'] or 0) * proportion if met['current_value'] else None
            gain         = (stake_value - invested) if stake_value is not None else None
            roi_pct      = (gain / invested * 100) if (gain is not None and invested) else None
            moic_val     = (stake_value / invested) if (stake_value and invested) else None

            data = [
                c['name'], c.get('sector', ''), c.get('country', ''),
                r.get('round_name', ''), r.get('date', ''),
                invested,
                r.get('pre_money_valuation'),
                r.get('post_money_valuation'),
                r.get('shares_received'),
                r.get('price_per_share'),
                r.get('ownership_pct'),
                c.get('current_valuation'),
                stake_value,
                gain,
                roi_pct,
                moic_val,
                r.get('status', ''),
            ]

            currency_cols = {6, 7, 8, 10, 12, 13, 14}
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                if col in currency_cols and isinstance(val, (int, float)):
                    cell.number_format = _CURRENCY
                elif col == 11 and isinstance(val, (int, float)):
                    cell.number_format = _PCT
                elif col == 15 and isinstance(val, (int, float)):
                    cell.number_format = '0.00"%"'
                elif col == 16 and isinstance(val, (int, float)):
                    cell.number_format = '0.00"×"'
                if col in (14, 15) and isinstance(val, (int, float)):
                    cell.fill = _GREEN_FILL if val >= 0 else _RED_FILL

            row_num += 1

    _autofit(ws)

    # ── Per-company sheets ────────────────────────────────────────────────────
    for c in companies:
        rounds = models.get_rounds(c['id'])
        if not rounds:
            continue

        met      = m.company_metrics(rounds, c.get('current_valuation'))
        ws2      = wb.create_sheet(c['name'][:31])

        ws2.cell(row=1, column=1, value=c['name']).font = Font(bold=True, size=13)
        summary_rows = [
            f"Total Invested:  {sym}{met['total_invested']:,.0f}",
            f"Current Value:   {sym}{met['current_value']:,.0f}" if met['current_value'] else "Current Value:  n/a (no valuation set)",
            f"Gain / Loss:     {sym}{met['gain']:,.0f} ({met['roi']:.1f}%)" if met['gain'] is not None else "Gain / Loss:    n/a",
            f"MOIC:            {met['moic']:.2f}×" if met['moic'] else "MOIC:           n/a",
            f"IRR:             {met['irr'] * 100:.1f}%" if met['irr'] else "IRR:            n/a",
        ]
        for i, line in enumerate(summary_rows, 2):
            ws2.cell(row=i, column=1, value=line)

        headers2 = ["Round", "Date", "Amount Invested", "Pre-Money", "Post-Money",
                    "Shares", "Price/Share", "Ownership %", "Status"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=8, column=col, value=h)
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL

        for i, r in enumerate(rounds, 9):
            ws2.cell(row=i, column=1, value=r.get('round_name', ''))
            ws2.cell(row=i, column=2, value=r.get('date', ''))
            ws2.cell(row=i, column=3, value=r.get('amount_invested')).number_format = _CURRENCY
            ws2.cell(row=i, column=4, value=r.get('pre_money_valuation'))
            ws2.cell(row=i, column=5, value=r.get('post_money_valuation'))
            ws2.cell(row=i, column=6, value=r.get('shares_received'))
            ws2.cell(row=i, column=7, value=r.get('price_per_share'))
            ws2.cell(row=i, column=8, value=r.get('ownership_pct'))
            ws2.cell(row=i, column=9, value=r.get('status', ''))

        ws2.freeze_panes = 'A9'
        _autofit(ws2)

    wb.save(path)


def _autofit(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or '')) for cell in col if cell.value is not None),
            default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)


# ── Family-fund format parser ─────────────────────────────────────────────────
# Exact column positions for "Sammanst_investeringar" spreadsheet (0-indexed):
#   Col  0  : Company name
#   Cols 1-9: Investment years 2018-2026 (amounts in TKR = thousands of SEK)
#   Col 10  : EXIT received — values >100 000 are in SEK (÷1000 → TKR); smaller = TKR
#   Col 11  : Totalt invested (TKR)
#   Col 12  : Värdering notes (Swedish text; may embed SEK stake amounts)
#   Col 13  : Sector (Område)
#   Col 16  : Exit year string  (e.g. "2026-IPO")
#   Col 18  : Target multiple  (numeric, e.g. 3.0)
#   Col 19  : Potential exit value (TKR)

_YEAR_COLS = {1: 2018, 2: 2019, 3: 2020, 4: 2021, 5: 2022,
              6: 2023, 7: 2024, 8: 2025, 9: 2026}
_C_EXIT      = 10
_C_TOTAL     = 11
_C_VAL_NOTES = 12
_C_SECTOR    = 13
_C_EXIT_YR   = 16
_C_MULTIPLE  = 18
_C_POTENTIAL = 19

_SKIP_PREFIXES = ('total', 'cash', 'number of', 'antal', 'summa ', 'snitt ')


def _col(row, idx):
    return row[idx] if idx < len(row) else None


def _f(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return None


def _extract_stake_tkr(notes: str) -> Optional[float]:
    """
    Find current stake value in TKR embedded in a Swedish valuation note.
    Looks for SEK amounts (space-separated thousands "23 947 307,16") ≥ 100 000 SEK
    and converts by ÷1000 to get TKR.
    """
    if not notes:
        return None
    # Swedish format with space thousands: "23 947 307" or "23 947 307,16"
    for m in re.findall(r'\b(\d{1,3}(?:\s\d{3})+(?:,\d{1,2})?)\b', notes):
        try:
            val = float(m.replace(' ', '').replace(',', '.'))
            if val >= 100_000:
                return round(val / 1000, 1)
        except ValueError:
            pass
    # Plain 7+ digit number without spaces
    for m in re.findall(r'\b(\d{7,})\b', notes):
        try:
            val = float(m)
            if val >= 100_000:
                return round(val / 1000, 1)
        except ValueError:
            pass
    return None


def _exit_to_tkr(raw) -> Optional[float]:
    """Convert EXIT column value to TKR. Values >100 000 are treated as SEK."""
    v = _f(raw)
    if v is None or v <= 0:
        return None
    return round(v / 1000, 1) if v > 100_000 else v


def parse_family_excel(path: str) -> List[Dict]:
    """
    Parse the family-fund spreadsheet (Sammanst_investeringar format).
    Uses hardcoded column positions — returns list of dicts for import_family_data().
    """
    wb   = openpyxl.load_workbook(path, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    results: List[Dict]  = []
    current_entity       = 'Portfolio A'
    is_exit_section      = False
    is_bankrupt_section  = False

    for row in rows:
        if not row or _col(row, 0) is None:
            continue

        raw_name = str(_col(row, 0)).strip()
        if not raw_name:
            continue

        lo = raw_name.lower()

        # ── Detect section headers ────────────────────────────────────────────
        if 'portfolio a' in lo and 'exit' not in lo:
            current_entity = 'Portfolio A'
            is_exit_section = False
            is_bankrupt_section = False
            continue
        if 'portfolio b' in lo and 'exit' not in lo:
            current_entity = 'Portfolio B'
            is_exit_section = False
            is_bankrupt_section = False
            continue
        if 'exits' in lo and 'portfolio a' in lo:
            current_entity = 'Portfolio A'
            is_exit_section = True
            is_bankrupt_section = False
            continue
        if 'exits' in lo and 'portfolio b' in lo:
            current_entity = 'Portfolio B'
            is_exit_section = True
            is_bankrupt_section = False
            continue
        if lo.startswith('konkurs'):
            current_entity = 'Portfolio A'
            is_bankrupt_section = True
            is_exit_section = False
            continue

        # ── Skip header/summary rows ──────────────────────────────────────────
        if any(lo.startswith(p) for p in _SKIP_PREFIXES):
            continue

        # ── Year investment amounts (TKR) ─────────────────────────────────────
        year_amounts: Dict[int, float] = {}
        for ci, yr in _YEAR_COLS.items():
            v = _f(_col(row, ci))
            if v and v > 0:
                year_amounts[yr] = v

        # ── Total invested (TKR) ──────────────────────────────────────────────
        total_tkr = _f(_col(row, _C_TOTAL))
        if not total_tkr:
            total_tkr = sum(year_amounts.values()) if year_amounts else 0.0
        total_tkr = total_tkr or 0.0

        if is_bankrupt_section:
            total_tkr = total_tkr or sum(year_amounts.values())

        # Skip rows with no investment data at all
        if not year_amounts and total_tkr == 0 and not is_exit_section:
            continue

        # ── Current / exit value ──────────────────────────────────────────────
        exit_tkr = _exit_to_tkr(_col(row, _C_EXIT))

        if is_bankrupt_section:
            current_val_tkr: Optional[float] = 0.0
        elif is_exit_section:
            current_val_tkr = exit_tkr          # exit received = final value
        else:
            current_val_tkr = _extract_stake_tkr(
                str(_col(row, _C_VAL_NOTES) or '')
            )

        # ── Other fields ──────────────────────────────────────────────────────
        val_notes = str(_col(row, _C_VAL_NOTES) or '').strip()
        sector    = str(_col(row, _C_SECTOR)    or '').strip()
        exit_year = str(_col(row, _C_EXIT_YR)   or '').strip()
        multiple  = _f(_col(row, _C_MULTIPLE))
        potential = _f(_col(row, _C_POTENTIAL))

        results.append({
            'name':              raw_name,
            'entity':            current_entity,
            'sector':            sector,
            'year_amounts':      year_amounts,       # {int year: float TKR}
            'total_invested':    total_tkr,           # TKR
            'current_valuation': current_val_tkr,     # TKR stake value, or None
            'multiple':          multiple,             # target ×
            'potential':         potential,            # expected exit TKR
            'exit_year':         exit_year,
            'is_exited':         is_exit_section,
            'is_bankrupt':       is_bankrupt_section,
            'valuation_notes':   val_notes,
        })

    return results


def export_family_excel(path: str) -> None:
    """
    Write all companies back to an Excel file in the family-fund format,
    so it can be re-imported or edited externally and re-synced.
    Columns match the hardcoded positions parse_family_excel() expects.
    """
    YEARS = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Blad1'

    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill(fill_type='solid', fgColor='1E293B')
    sec_fill  = PatternFill(fill_type='solid', fgColor='2563EB')
    sec_font  = Font(bold=True, color='FFFFFF', size=10)
    tot_fill  = PatternFill(fill_type='solid', fgColor='EFF6FF')

    # ── Row 1: human-readable header ─────────────────────────────────────────
    headers = (
        ['Bolag'] + [str(y) for y in YEARS] +
        ['EXIT', 'Totalt', 'Värdering per 31 Dec 2025',
         'Sektor', 'Kommentarer', '', 'Exit år', '', 'Multipel', 'Potential']
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    ws.freeze_panes = 'A2'

    # Group companies by entity, preserving Portfolio A → Portfolio B order
    entities_order = []
    entities_map: Dict[str, list] = {}
    for c in models.get_all_companies():
        e = c.get('entity') or 'Other'
        if e not in entities_map:
            entities_map[e] = []
            entities_order.append(e)
        entities_map[e].append(c)

    row_num = 2
    for entity in entities_order:
        cos = entities_map[entity]

        # Separate active, exited, bankrupt within this entity
        active   = [c for c in cos if 'status: exited'   not in (c.get('notes') or '').lower()
                    and 'bankrupt' not in (c.get('notes') or '').lower()]
        exited   = [c for c in cos if 'status: exited'   in (c.get('notes') or '').lower()]
        bankrupt = [c for c in cos if 'bankrupt'         in (c.get('notes') or '').lower()]

        def _write_section_header(title):
            nonlocal row_num
            cell = ws.cell(row=row_num, column=1, value=title)
            cell.font = sec_font
            cell.fill = sec_fill
            row_num += 1

        def _write_company(c):
            nonlocal row_num
            rounds = models.get_rounds(c['id'])
            yr_amt: Dict[int, float] = {}
            for r in rounds:
                yr_str = (r.get('date') or '')[:4]
                try:
                    yr = int(yr_str)
                    if yr in YEARS and r.get('amount_invested'):
                        yr_amt[yr] = yr_amt.get(yr, 0) + r['amount_invested']
                except ValueError:
                    pass

            notes = c.get('notes') or ''
            exit_yr = ''
            multiple = None
            potential = None
            for line in notes.splitlines():
                if line.startswith('Exit year:'):
                    exit_yr = line.split(':', 1)[1].strip()
                elif line.startswith('Target multiple:'):
                    try:
                        multiple = float(line.split(':', 1)[1].replace('×', '').strip())
                    except ValueError:
                        pass
                elif line.startswith('Potential at exit:'):
                    try:
                        potential = float(line.split(':', 1)[1].replace('TKR', '').replace(',', '').strip())
                    except ValueError:
                        pass

            total = sum(yr_amt.values()) or (c.get('current_valuation') or 0)
            ws.cell(row=row_num, column=1, value=c['name'])
            for ci, yr in enumerate(YEARS, 2):
                val = yr_amt.get(yr)
                if val:
                    ws.cell(row=row_num, column=ci, value=round(val, 2))
            # col 11 = EXIT (skip), col 12 = Totalt
            ws.cell(row=row_num, column=12, value=round(total, 2) if total else None)
            # col 13 = Värdering
            if c.get('current_valuation') is not None:
                ws.cell(row=row_num, column=13, value=round(c['current_valuation'], 1))
            ws.cell(row=row_num, column=14, value=c.get('sector') or '')
            ws.cell(row=row_num, column=17, value=exit_yr or None)
            ws.cell(row=row_num, column=19, value=multiple)
            ws.cell(row=row_num, column=20, value=potential)
            row_num += 1

        def _write_total(label, cos_list):
            nonlocal row_num
            totals = [0.0] * len(YEARS)
            grand  = 0.0
            for c in cos_list:
                for r in models.get_rounds(c['id']):
                    yr_str = (r.get('date') or '')[:4]
                    try:
                        yr_i = YEARS.index(int(yr_str))
                        totals[yr_i] += r.get('amount_invested') or 0
                        grand         += r.get('amount_invested') or 0
                    except (ValueError, IndexError):
                        pass
            ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            for ci, v in enumerate(totals, 2):
                if v:
                    cell = ws.cell(row=row_num, column=ci, value=round(v, 2))
                    cell.fill = tot_fill
            ws.cell(row=row_num, column=12, value=round(grand, 2)).fill = tot_fill
            row_num += 1

        _write_section_header(entity)
        for c in active:
            _write_company(c)
        _write_total(f"Total {entity}", active)

        if exited:
            _write_section_header(f"EXITS {entity}")
            for c in exited:
                _write_company(c)

        if bankrupt:
            _write_section_header("Konkurser")
            for c in bankrupt:
                _write_company(c)

        row_num += 1   # blank row between entities

    # Column widths
    ws.column_dimensions['A'].width = 30
    for ci in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 10
    ws.column_dimensions['M'].width = 22

    wb.save(path)


def import_family_data(companies: List[Dict]) -> int:
    """
    Write parsed family-format data into the DB.
    Returns count of NEW companies created.
    Duplicate detection is (name, entity) — same name in different entities = different company.
    """
    models.set_setting('currency',      'TKR')
    models.set_setting('currency_name', 'TKR')

    count = 0
    # Key: (name_lower, entity_lower) so same company in different portfolios stays separate
    all_existing: Dict[tuple, dict] = {
        (c['name'].lower(), (c.get('entity') or '').lower()): c
        for c in models.get_all_companies()
    }

    for co in companies:
        entity = co.get('entity', '').strip()

        # Build notes string
        notes_parts: List[str] = []
        if co.get('is_exited'):
            notes_parts.append('Status: Exited')
        if co.get('is_bankrupt'):
            notes_parts.append('Status: Bankrupt (written off)')
        if co.get('valuation_notes'):
            notes_parts.append(f"Valuation note: {co['valuation_notes']}")
        if co.get('exit_year'):
            notes_parts.append(f"Exit year: {co['exit_year']}")
        if co.get('multiple'):
            notes_parts.append(f"Target multiple: {co['multiple']}×")
        if co.get('potential'):
            notes_parts.append(f"Potential at exit: {co['potential']:,.0f} TKR")
        notes = '\n'.join(notes_parts)

        key = (co['name'].lower(), entity.lower())
        existing = all_existing.get(key)

        if existing:
            cid = existing['id']
            upd: dict = {'notes': notes}
            if co.get('sector'):
                upd['sector'] = co['sector']
            if co.get('current_valuation') is not None:
                upd['current_valuation'] = co['current_valuation']
            models.update_company(cid, **upd)
        else:
            first_year = min(co['year_amounts'].keys()) if co['year_amounts'] else None
            cid = models.add_company(
                name=co['name'],
                entity=entity,
                sector=co.get('sector', ''),
                country='',
                first_investment_date=f"{first_year}-01-01" if first_year else '',
                current_valuation=co.get('current_valuation'),
                notes=notes,
            )
            all_existing[key] = {'id': cid, 'name': co['name'], 'entity': entity}
            count += 1

        # Clear and recreate rounds from year amounts (ensures re-import stays accurate)
        models.clear_rounds(cid)
        for yr in sorted(co['year_amounts']):
            models.add_round(
                company_id=cid,
                round_name=str(yr),
                date=f"{yr}-07-01",
                amount_invested=co['year_amounts'][yr],
                ownership_pct=100.0,
                status='Closed',
            )

        # Fallback: no year amounts but has total (e.g. shares received as distribution)
        total_tkr = co.get('total_invested', 0) or 0
        if not co['year_amounts'] and total_tkr > 0:
            models.add_round(
                company_id=cid,
                round_name='Investment',
                date='2023-07-01',
                amount_invested=total_tkr,
                ownership_pct=100.0,
                status='Closed',
            )

    return count
